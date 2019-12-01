#coding=utf-8
from openpyxl import Workbook
#导入相关模块和类
wb=Workbook()   
#创建文件对象
ws=wb.active
# 获取第一个sheet
ws['A1']=42
# 写入数据到单元格中
ws['B1']='come on baby'
ws.append([1,2,3])
wb.save("D:\\test1201.xlsx")
print('oyes!!')